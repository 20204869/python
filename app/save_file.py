# coding=utf-8
'''
数据写文件
'''

import pandas as pd
from openpyxl import load_workbook

#数据结果写入excel（覆盖sheet内容）
def write_excel(result_data, path_file_name):
    writer = pd.ExcelWriter(path_file_name, engine='openpyxl')
    book = load_workbook(path_file_name)
    writer.book = book

    # 解决excel sheet 数据部分覆盖问题
    idx = book.sheetnames.index('test')
    book.remove(book.worksheets[idx])
    book.create_sheet('test', idx)

    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # 数据结果格式处理并写入excel
    data = result_data.apply(pd.to_numeric, errors='ignore')
    data.to_excel(writer, sheet_name="test", index=None)
    writer.save()


#数据追加写入excel
def write_append_excel(result_data, path_file_name):
    # 读取excel原文件信息
    old_result = pd.DataFrame(pd.read_excel(path_file_name, sheet_name='sheet'))
    writer = pd.ExcelWriter(path_file_name, engine='openpyxl')
    book = load_workbook(path_file_name)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # 获取原数据的行数
    df_rows = old_result.shape[0]
    data = result_data.toPandas()
    data.to_excel(writer, "sheet", startrow=df_rows + 1, index=None)
    writer.save()


#多结果数据写excel多sheet
#PS: 多参数传入方式：*result_data  多参传入类型元组[tuple]
def more_sheets(result_data, path_file_name):
    writer = pd.ExcelWriter(path_file_name)
    data = result_data.toPandas()
    # 设置多数据结果
    d1 = data.copy()
    d2 = data.copy()
    d3 = data.copy()
    d4 = data.copy()

    d1.to_excel(writer, "d1", index=None)
    d2.to_excel(writer, "d2", index=None)
    d3.to_excel(writer, "d3", index=None)
    d4.to_excel(writer, "d4", index=None)
    writer.save()


#多结果写多sheet
def result2excel(*result_data, file_path_name):
    writer = pd.ExcelWriter(file_path_name, engine='openpyxl')
    book = load_workbook(file_path_name)
    writer.book = book

    # 设置sheet名称
    idx1 = book.sheetnames.index('test')
    book.remove(book.worksheets[idx1])
    book.create_sheet('test', idx1)

    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # 查询结果
    data_1 = result_data[0].apply(pd.to_numeric, errors='ignore')
    data_1.to_excel(writer, sheet_name="test", index=None)

    # 指定从哪一列开始写入
    data_2 = result_data[1].apply(pd.to_numeric, errors='ignore')
    data_2.to_excel(writer, sheet_name="test", startcol=5, index=None)

    writer.save()


#数据写csv PS : 表数据量过大会出现内存溢出OOM
def result2csv(data,path_file_name):
    user_car_data=data.toPandas()
    user_car_data.to_csv(path_file_name,index=False,mode='append')
